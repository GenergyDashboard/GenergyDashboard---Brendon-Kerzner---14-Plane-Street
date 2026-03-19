"""
14 Plane Street – Data Processor
Reads VRM kWh XLSX exports → outputs Nautica-compatible dashboard JSON.

Usage:
    python process_data.py --today downloads/14PlanestreetJBay_kwh_20260317.xlsx \
                           --lifetime data/Monthly_Data_Excluding_today.xlsx \
                           --output data/dashboard_data.json \
                           --append-yesterday downloads/14PlanestreetJBay_kwh_20260316.xlsx
"""
import argparse, json, os
from collections import defaultdict
from datetime import datetime, timedelta
import openpyxl

FIELDS = ["solar_yield","grid_to_batt","grid_to_cons","pv_to_batt","pv_to_grid",
          "pv_to_cons","batt_to_cons","batt_to_grid","genset_to_cons","genset_to_batt"]

VRM_HEADERS = ['timestamp','Solar Yield (delta)','Grid to battery','Grid to consumers',
               'PV to battery','PV to grid','PV to consumers','Battery to consumers',
               'Battery to grid','Genset to consumers','Genset to battery','Gas']
VRM_UNITS = ['Africa/Johannesburg (+02:00)','kWh','kWh','kWh','kWh','kWh','kWh','kWh','kWh','kWh','kWh','m3']

def append_to_lifetime(daily_xlsx, lifetime_xlsx):
    if not os.path.exists(daily_xlsx):
        print(f"  [append] Not found: {daily_xlsx}"); return
    if os.path.exists(lifetime_xlsx):
        wb = openpyxl.load_workbook(lifetime_xlsx); ws = wb.active
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "VRM kWh report"
        ws.append([None]*12); ws.append(VRM_HEADERS); ws.append(VRM_UNITS)
    existing = set()
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ts = str(row[0]).strip() if row[0] else None
        if ts: existing.add(ts)
    wb_d = openpyxl.load_workbook(daily_xlsx, data_only=True); ws_d = wb_d.active
    added = 0
    for row in ws_d.iter_rows(min_row=4, max_row=ws_d.max_row, values_only=True):
        ts = str(row[0]).strip() if row[0] else None
        if not ts or ts in existing: continue
        ws.append(list(row)); added += 1
    wb.save(lifetime_xlsx)
    print(f"  [append] +{added} rows (was {len(existing)}, now {len(existing)+added})")

def parse_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True); ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ts = str(row[0]).strip() if row[0] else None
        if not ts: continue
        entry = {"ts": ts}
        for i, f in enumerate(FIELDS): entry[f] = float(row[i+1] or 0)
        rows.append(entry)
    return rows

def derive(d):
    d["pv_total"] = d["pv_to_batt"] + d["pv_to_grid"] + d["pv_to_cons"]
    d["load"] = d["grid_to_cons"] + d["pv_to_cons"] + d["batt_to_cons"] + d["genset_to_cons"]
    d["grid_import"] = d["grid_to_cons"] + d["grid_to_batt"]
    d["export"] = d["pv_to_grid"] + d["batt_to_grid"]
    d["self_consumption"] = d["pv_to_cons"]
    return d

def aggregate(rows, key_fn):
    agg = defaultdict(lambda: {k:0.0 for k in FIELDS})
    for r in rows:
        k = key_fn(r["ts"])
        for f in FIELDS: agg[k][f] += r[f]
    result = {}
    for k in sorted(agg.keys()):
        entry = dict(agg[k]); derive(entry); result[k] = entry
    return result

def make_period_data(d):
    """Convert aggregated dict to Nautica-format period data block."""
    return {
        "PV Yield (kWh)": round(d.get("pv_total",0), 2),
        "Self-consumption (kWh)": round(d.get("pv_to_cons",0), 2),
        "Export (kWh)": round(d.get("export",0), 2),
        "Consumption (kWh)": round(d.get("load",0), 2),
        "Import (kWh)": round(d.get("grid_import",0), 2),
        "Battery to Load (kWh)": round(d.get("batt_to_cons",0), 2),
        "PV to Battery (kWh)": round(d.get("pv_to_batt",0), 2),
        "Battery to Grid (kWh)": round(d.get("batt_to_grid",0), 2),
        "Grid to Battery (kWh)": round(d.get("grid_to_batt",0), 2),
    }

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--today", required=True)
    p.add_argument("--lifetime", required=True)
    p.add_argument("--output", default="data/dashboard_data.json")
    p.add_argument("--append-yesterday", default=None)
    args = p.parse_args()

    if args.append_yesterday:
        print(f"Appending yesterday: {args.append_yesterday}")
        append_to_lifetime(args.append_yesterday, args.lifetime)

    print(f"Reading today: {args.today}")
    today_raw = parse_xlsx(args.today)
    print(f"  → {len(today_raw)} rows")

    print(f"Reading lifetime: {args.lifetime}")
    lifetime_raw = parse_xlsx(args.lifetime)
    print(f"  → {len(lifetime_raw)} rows")

    all_raw = lifetime_raw + today_raw
    today_date = today_raw[-1]["ts"][:10] if today_raw else datetime.now().strftime("%Y-%m-%d")

    # Aggregate at different levels
    hourly_all = aggregate(all_raw, lambda ts: ts[:13]+":00:00")
    daily_all = aggregate(all_raw, lambda ts: ts[:10])
    monthly_all = aggregate(all_raw, lambda ts: ts[:7])

    # Today's hourly
    today_hourly = aggregate(today_raw, lambda ts: ts[:13]+":00:00")
    pv_h = [0.0]*24; ld_h = [0.0]*24; gi_h = [0.0]*24; bc_h = [0.0]*24; ex_h = [0.0]*24
    pvc_h = [0.0]*24; pvb_h = [0.0]*24; gc_h = [0.0]*24; gb_h = [0.0]*24; bg_h = [0.0]*24
    for k, v in today_hourly.items():
        h = int(k[11:13])
        pv_h[h] = round(v["pv_total"],3); ld_h[h] = round(v["load"],3)
        gi_h[h] = round(v["grid_import"],3); bc_h[h] = round(v["batt_to_cons"],3)
        ex_h[h] = round(v["export"],3)
        pvc_h[h] = round(v.get("pv_to_cons",0),3); pvb_h[h] = round(v.get("pv_to_batt",0),3)
        gc_h[h] = round(v.get("grid_to_cons",0),3); gb_h[h] = round(v.get("grid_to_batt",0),3)
        bg_h[h] = round(v.get("batt_to_grid",0),3)
    current_hour = max((int(k[11:13]) for k in today_hourly.keys()), default=0)

    # Average hourly profiles (MTD for current month)
    cur_month = today_date[:7]
    month_days = [d for d in sorted(daily_all.keys()) if d[:7] == cur_month]
    avg_load = [0.0]*24; avg_grid = [0.0]*24; avg_pv = [0.0]*24
    if month_days:
        for d_date in month_days:
            for k, v in hourly_all.items():
                if k[:10] == d_date:
                    h = int(k[11:13])
                    avg_load[h] += v["load"]; avg_grid[h] += v["grid_import"]; avg_pv[h] += v["pv_total"]
        n = len(month_days)
        avg_load = [round(x/n,3) for x in avg_load]
        avg_grid = [round(x/n,3) for x in avg_grid]
        avg_pv = [round(x/n,3) for x in avg_pv]

    # Yesterday
    from datetime import timedelta
    yesterday_str = (datetime.strptime(today_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    yesterday_data = daily_all.get(yesterday_str, {f:0 for f in FIELDS})
    if "pv_total" not in yesterday_data: derive(yesterday_data)

    # Today/Month/Lifetime aggregated
    today_agg = daily_all.get(today_date, {f:0 for f in FIELDS})
    if "pv_total" not in today_agg: derive(today_agg)

    month_agg = {f:0.0 for f in FIELDS}
    for d, v in daily_all.items():
        if d[:7] == cur_month:
            for f in FIELDS: month_agg[f] += v[f]
    derive(month_agg)

    life_agg = {f:0.0 for f in FIELDS}
    for v in daily_all.values():
        for f in FIELDS: life_agg[f] += v[f]
    derive(life_agg)

    # Daily history (for daily tab date picker)
    daily_hist = {}
    for d_date, d_data in sorted(daily_all.items()):
        rec = {
            "pv": round(d_data["pv_total"],2),
            "consumption": round(d_data["load"],2),
            "import": round(d_data["grid_import"],2),
            "export": round(d_data["export"],2),
            "self_consumption": round(d_data["pv_to_cons"],2),
            "batt_to_cons": round(d_data["batt_to_cons"],2),
            "pv_to_batt": round(d_data["pv_to_batt"],2),
            "batt_to_grid": round(d_data["batt_to_grid"],2),
            "grid_to_batt": round(d_data["grid_to_batt"],2),
        }
        # Hourly breakdown for this day
        day_hourly = {"pv":[0]*24, "load":[0]*24, "grid":[0]*24, "batt":[0]*24, "export":[0]*24,
                      "pv_to_cons":[0]*24, "pv_to_batt":[0]*24, "grid_to_cons":[0]*24,
                      "grid_to_batt":[0]*24, "batt_to_grid":[0]*24}
        for k, v in hourly_all.items():
            if k[:10] == d_date:
                h = int(k[11:13])
                day_hourly["pv"][h] = round(v["pv_total"],3)
                day_hourly["load"][h] = round(v["load"],3)
                day_hourly["grid"][h] = round(v["grid_import"],3)
                day_hourly["batt"][h] = round(v["batt_to_cons"],3)
                day_hourly["export"][h] = round(v["export"],3)
                day_hourly["pv_to_cons"][h] = round(v.get("pv_to_cons",0),3)
                day_hourly["pv_to_batt"][h] = round(v.get("pv_to_batt",0),3)
                day_hourly["grid_to_cons"][h] = round(v.get("grid_to_cons",0),3)
                day_hourly["grid_to_batt"][h] = round(v.get("grid_to_batt",0),3)
                day_hourly["batt_to_grid"][h] = round(v.get("batt_to_grid",0),3)
        rec["hourly"] = day_hourly
        daily_hist[d_date] = rec

    # Lifetime monthly breakdown (for lifetime tab)
    lifetime = {}
    for m_key, m_data in sorted(monthly_all.items()):
        yr, mo = m_key.split("-")
        if yr not in lifetime: lifetime[yr] = {}
        lifetime[yr][str(int(mo))] = {
            "pv": round(m_data["pv_total"],2),
            "consumption": round(m_data["load"],2),
            "import": round(m_data["grid_import"],2),
            "export": round(m_data["export"],2),
            "self_consumption": round(m_data["pv_to_cons"],2),
            "batt_to_cons": round(m_data["batt_to_cons"],2),
            "pv_to_batt": round(m_data["pv_to_batt"],2),
        }

    # Build output in Nautica-compatible format
    output = {
        "site_name": "14 Plane Street",
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "today": {"data": make_period_data(today_agg), "date": today_date},
        "yesterday": {"data": make_period_data(yesterday_data)},
        "current_month": {"data": make_period_data(month_agg)},
        "all_time_totals": make_period_data(life_agg),
        "hourly": {
            "pv": pv_h, "today": pv_h,
            "load": ld_h, "grid": gi_h,
            "batt": bc_h, "export": ex_h,
            "pv_to_cons": pvc_h, "pv_to_batt": pvb_h,
            "grid_to_cons": gc_h, "grid_to_batt": gb_h,
            "batt_to_grid": bg_h,
            "current_hour": current_hour,
            "avg_load": avg_load, "avg_grid": avg_grid, "avg_pv": avg_pv,
        },
        "lifetime": lifetime,
    }

    # Add monthly aggregates keyed by "YYYY-MM" (required by dashboard)
    monthly_keyed = {}
    for m_key, m_data in sorted(monthly_all.items()):
        monthly_keyed[m_key] = make_period_data(m_data)
    output["monthly"] = monthly_keyed

    # Write both files
    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(output, f, separators=(",", ":"))

    # Also write daily_history.json (separate file like Nautica)
    hist_path = os.path.join(os.path.dirname(args.output), "daily_history.json")
    with open(hist_path, "w") as f:
        json.dump(daily_hist, f, separators=(",", ":"))

    sz1 = os.path.getsize(args.output)/1024
    sz2 = os.path.getsize(hist_path)/1024
    print(f"\nWrote {args.output} ({sz1:.0f} KB)")
    print(f"Wrote {hist_path} ({sz2:.0f} KB)")
    print(f"  Days: {len(daily_all)}, Months: {len(monthly_all)}")
    print(f"  PV total: {life_agg['pv_total']:.1f} kWh, Load: {life_agg['load']:.1f} kWh")

if __name__ == "__main__":
    main()
